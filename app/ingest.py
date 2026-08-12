"""
File ingestion - streaming and memory-bounded.

Why this is written the way it is
---------------------------------
Version 1 looped with df.iterrows() and committed once at the end. That
degraded badly past ~13k rows.

Version 2 fixed speed with vectorised pandas + Postgres COPY, but still
read the entire file into memory twice: once as raw bytes in the web
request, then again as a DataFrame. A 200MB CSV therefore needed roughly
600MB-1GB of RAM to import, which is how a small container gets
OOM-killed mid-upload.

This version bounds memory regardless of file size:

* The web layer streams the upload to a temp file on disk in 1MB pieces
  and hands ingest a PATH, not bytes. Raw file content is never fully
  resident in RAM.
* Fact files are read in row chunks (default 50k) and each chunk is
  converted and COPY'd straight into Postgres, then released. Peak memory
  is a function of CHUNK_ROWS, not of file size. A 200MB file and a 2GB
  file now use the same amount of RAM.
* Excel is converted to CSV on disk first using openpyxl's read_only
  streaming reader, so .xlsx gets the same bounded-memory path instead of
  the 10-20x in-memory XML blowup that pd.read_excel() incurs.
* Date deletion happens ONCE up front (a cheap scan of just the date
  column) rather than per chunk - otherwise chunk 2 would delete the rows
  chunk 1 just inserted for the same date.
* Per-chunk stats are accumulated as counters and turned into notes at
  the end, so a message reads "233,000 blank rows" once rather than once
  per chunk with the wrong number.

Loading remains IDEMPOTENT per (table, date): re-uploading a file for a
date clears that date first, so overlapping historical exports are safe.

Masters (store_master, product_master) are small and full-replace, so
they are still read whole - chunking them would break cross-chunk
de-duplication for no benefit.
"""
import io
import os
import csv
import tempfile
from pathlib import Path

import pandas as pd
from sqlalchemy import delete

from .models import (SessionLocal, engine, DimStore, DimProduct, FactDispatch,
                     FactStoreReceiving, FactWarehouseReceiving, FactReject,
                     FactRoute, FactIndent, UploadLog,
                     canon_category, FSN_PREFIX_CATEGORY)

# Rows held in memory at once. Lower this if the container is very small;
# raise it for faster imports on a bigger box. Memory scales with this
# number, NOT with the size of the file being imported.
CHUNK_ROWS = int(os.getenv("INGEST_CHUNK_ROWS", "50000"))

# Rows read to sniff the column layout before committing to a strategy.
SAMPLE_ROWS = 200

SIGNATURES = {
    "batching": {"po number", "cutoff datetime", "fsn", "store_id",
                 "total picked qty"},
    "store_receiving": {"warehouse id", "invoice id", "fsn",
                        "expected quantity", "received quantity"},
    "wh_receiving": {"fsn", "product", "brand", "category"},
    "rejects": {"fsn", "product", "qty", "reason"},
    "store_master": {"warehouse name", "facility site code / wh"},
    "product_master": {"brand", "category", "ean", "title", "mrp"},
    "indent": {"po qty", "vertical", "title"},
    "route": {"vehicle no", "store no"},
}

MASTER_TYPES = {"store_master", "product_master"}


def _norm_cols(df):
    return {str(c).strip().lower().replace("\n", " ") for c in df.columns}


def detect_type(df):
    cols = _norm_cols(df)
    has = lambda *needles: any(all(x in c for x in needles) for c in cols)

    # RECEIVING and STORE_REJECTS both carry fsn/product/brand/category;
    # only rejects has 'reason', only indent has 'vertical'.
    if has("reason") and has("qty") and not has("expected", "quantity"):
        return "rejects", 1.0
    if has("vertical") and has("po", "qty"):
        return "indent", 1.0
    if has("po", "qty") and has("received") and not has("cutoff") and not has("warehouse", "id"):
        return "wh_receiving", 1.0

    scores = {}
    for name, sig in SIGNATURES.items():
        hits = sum(1 for s in sig if any(s in c for c in cols))
        scores[name] = hits / len(sig)
    best, score = max(scores.items(), key=lambda kv: kv[1])
    return (best, score) if score >= 0.6 else (None, score)


def _col(df, *needles):
    for c in df.columns:
        n = str(c).strip().lower().replace("\n", " ")
        if all(x in n for x in needles):
            return c
    return None


# ---------------- file access (bounded memory) ----------------
def excel_to_csv(src_path, dst_path):
    """
    Stream an .xlsx/.xlsm to CSV without building the whole workbook in
    memory. openpyxl's read_only mode yields rows lazily; pd.read_excel()
    by contrast parses the entire sheet XML up front, which is what makes
    a large workbook demand multiple GB.

    Dates survive as string timestamps, so the Excel-mangled-QTY recovery
    downstream still works.
    """
    from openpyxl import load_workbook
    wb = load_workbook(src_path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        with open(dst_path, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            for row in ws.iter_rows(values_only=True):
                w.writerow(["" if v is None else v for v in row])
    finally:
        wb.close()
    return dst_path


def to_csv_path(filename, path, tmpdir):
    """Return (csv_path, was_converted) for any supported input."""
    low = filename.lower()
    if low.endswith((".xlsx", ".xlsm")):
        out = os.path.join(tmpdir, "converted.csv")
        return excel_to_csv(path, out), True
    if low.endswith(".xls"):
        # Legacy .xls has no streaming reader; these files are small in
        # practice, so a whole-file read is acceptable here.
        out = os.path.join(tmpdir, "converted.csv")
        pd.read_excel(path).to_csv(out, index=False)
        return out, True
    return path, False


def detect_encoding(path):
    for enc in ("utf-8", "latin-1"):
        try:
            pd.read_csv(path, nrows=5, encoding=enc)
            return enc
        except UnicodeDecodeError:
            continue
        except Exception:
            # A parse error is not an encoding error - the encoding is fine.
            return enc
    return "utf-8"


def read_sample(path, encoding, nrows=SAMPLE_ROWS):
    return pd.read_csv(path, encoding=encoding, encoding_errors="replace",
                       nrows=nrows, low_memory=False)


def iter_chunks(path, encoding, chunk_rows=CHUNK_ROWS, usecols=None):
    return pd.read_csv(path, encoding=encoding, encoding_errors="replace",
                       chunksize=chunk_rows, usecols=usecols, low_memory=False)


def read_whole(path, encoding):
    return pd.read_csv(path, encoding=encoding, encoding_errors="replace",
                       low_memory=False)


# ---------------- vectorised converters ----------------
def v_int(s):
    """Whole-column numeric coercion; non-numeric and blank become 0."""
    return pd.to_numeric(s, errors="coerce").fillna(0).astype("int64")


def v_int_null(s):
    """As above but preserves NULL, for genuinely optional numbers."""
    return pd.to_numeric(s, errors="coerce").astype("Int64")


def v_date(s):
    return pd.to_datetime(s, errors="coerce", dayfirst=True)


def v_str(s, lower=False, maxlen=None):
    out = s.astype("string").fillna("").str.strip()
    if lower:
        out = out.str.lower()
    if maxlen:
        out = out.str.slice(0, maxlen)
    return out


def v_category_from_fsn(fsn_series):
    """Vectorised FSN-prefix category fallback. F&V has no product-master
    entry, so without this it silently vanishes from every breakdown."""
    return (fsn_series.str.slice(0, 3).str.upper()
            .map(FSN_PREFIX_CATEGORY).fillna("Unknown"))


def fix_excel_date_qty_vec(s):
    """
    Excel converts quantities to dates when the cell is date-formatted -
    typing 5 becomes 1900-01-05. Recover the integer from the 1900 serial.
    Returns (values, was_corrupted_flags).
    """
    numeric = pd.to_numeric(s, errors="coerce")
    as_dt = pd.to_datetime(s, errors="coerce")
    is_1900 = as_dt.dt.year.eq(1900).fillna(False)
    recovered = (as_dt - pd.Timestamp("1899-12-31")).dt.days
    vals = numeric.where(~is_1900, recovered).fillna(0).astype("int64")
    corrupted = (is_1900 | (numeric.isna() & s.notna())).fillna(False)
    return vals, corrupted


# ---------------- bulk load ----------------
def copy_into(df, table_name, columns, progress=None):
    """
    Stream a DataFrame into Postgres with COPY, committing as it goes.
    Falls back to chunked to_sql on SQLite, which has no COPY.
    """
    total = len(df)
    if total == 0:
        return 0

    is_pg = engine.dialect.name == "postgresql"
    done = 0

    for start in range(0, total, CHUNK_ROWS):
        chunk = df.iloc[start:start + CHUNK_ROWS]

        if is_pg:
            buf = io.StringIO()
            chunk.to_csv(buf, index=False, header=False, na_rep="\\N",
                         quoting=csv.QUOTE_MINIMAL)
            buf.seek(0)
            raw = engine.raw_connection()
            try:
                cur = raw.cursor()
                cur.copy_expert(
                    f"COPY {table_name} ({', '.join(columns)}) "
                    f"FROM STDIN WITH (FORMAT CSV, NULL '\\N')", buf)
                raw.commit()
            finally:
                raw.close()
            buf.close()
        else:
            chunk.to_sql(table_name, engine, if_exists="append", index=False,
                         method="multi", chunksize=500)

        done += len(chunk)
        if progress:
            progress(done, total)

    return done


# ---------------- per-import caches ----------------
def _valid_stores(ctx, db):
    """Store list is queried once per import, not once per chunk."""
    if "stores" not in ctx:
        ctx["stores"] = {s.warehouse_id for s in db.query(DimStore).all()}
    return ctx["stores"]


def _product_master(ctx, db):
    if "pm" not in ctx:
        rows = db.query(DimProduct.fsn, DimProduct.category,
                        DimProduct.brand).all()
        ctx["pm"] = (pd.DataFrame(rows, columns=["fsn", "pm_category", "pm_brand"])
                     if rows else
                     pd.DataFrame(columns=["fsn", "pm_category", "pm_brand"]))
    return ctx["pm"]


def _attach_cat_brand(out, ctx, db):
    """Left-join product master, then fall back to the FSN prefix for
    category - as a whole-frame merge, not per-row lookups."""
    pm = _product_master(ctx, db)
    merged = out.merge(pm, on="fsn", how="left")
    merged["category"] = merged["pm_category"].fillna(v_category_from_fsn(merged["fsn"]))
    merged["brand"] = merged["pm_brand"]
    return merged.drop(columns=["pm_category", "pm_brand"])


def _replace_dates(db, model, date_col, dates):
    if dates:
        db.execute(delete(model).where(date_col.in_(dates)))
        db.commit()


def _dates_of(dt_series):
    return sorted({d for d in dt_series.dropna().dt.date.unique()})


# --------------------------------------------------------------------------
# Loaders -> (rows_loaded, rows_dropped, dates, stats)
#
# stats is a dict of counters. The driver sums them across chunks and
# summarise() turns the totals into operator-facing notes once, at the end.
# --------------------------------------------------------------------------
def load_store_master(db, df, ctx, progress=None, replace=True):
    ser = _col(df, "serial")
    out = pd.DataFrame({
        "warehouse_id": v_str(df[_col(df, "facility")], lower=True),
        "warehouse_name": v_str(df[_col(df, "warehouse name")]),
        "wh_serial_no": v_int(df[ser]) if ser else 0,
    })
    out["city_code"] = out["warehouse_id"].str.split("_").str[0]
    if replace:
        db.query(DimStore).delete()
        db.commit()
        ctx.pop("stores", None)
    cols = ["warehouse_id", "warehouse_name", "wh_serial_no", "city_code"]
    n = copy_into(out[cols], "dim_store", cols, progress)
    return n, 0, [], {}


def load_product_master(db, df, ctx, progress=None, replace=True):
    n_in = len(df)
    out = pd.DataFrame({
        "fsn": v_str(df[_col(df, "fsn")]),
        "ean": v_str(df[_col(df, "ean")], maxlen=20),
        "brand": v_str(df[_col(df, "brand")]),
        "category": df[_col(df, "category")].map(canon_category),
        "title": v_str(df[_col(df, "title")]),
        "mrp": pd.to_numeric(df[_col(df, "mrp")], errors="coerce"),
        "price": pd.to_numeric(df[_col(df, "price")], errors="coerce"),
    })
    out = out[out["fsn"] != ""].drop_duplicates(subset="fsn", keep="first")
    dropped = n_in - len(out)
    if replace:
        db.query(DimProduct).delete()
        db.commit()
        ctx.pop("pm", None)
    cols = ["fsn", "ean", "brand", "category", "title", "mrp", "price"]
    n = copy_into(out[cols], "dim_product", cols, progress)
    return n, dropped, [], {"dupes": dropped}


def load_batching(db, df, ctx, progress=None, replace=True):
    cutoff = _col(df, "cutoff")
    dt = v_date(df[cutoff])
    dates = _dates_of(dt)
    if replace:
        _replace_dates(db, FactDispatch, FactDispatch.dispatch_date, dates)

    out = pd.DataFrame({
        "dispatch_date": dt.dt.date,
        "cutoff_datetime": v_str(df[cutoff]),
        "po_number": v_str(df[_col(df, "po number")]),
        "fsn": v_str(df[_col(df, "fsn")]),
        "product_title": v_str(df[_col(df, "product title")]),
        "warehouse_id": v_str(df[_col(df, "store_id") or _col(df, "store", "id")],
                              lower=True),
        "expected_qty": v_int(df[_col(df, "expected qty")]),
        "picked_qty": v_int(df[_col(df, "picked qty")]),
        "shortage_qty": v_int(df[_col(df, "shortage qty")]),
        "pending_qty": v_int(df[_col(df, "pending qty")]),
        "status": v_str(df[_col(df, "status")]),
        "picked_by": v_str(df[_col(df, "picked by")]),
        "picked_at": v_str(df[_col(df, "picked at")]),
    })
    out = _attach_cat_brand(out, ctx, db)
    cols = ["dispatch_date", "cutoff_datetime", "po_number", "fsn", "product_title",
            "warehouse_id", "expected_qty", "picked_qty", "shortage_qty",
            "pending_qty", "status", "picked_by", "picked_at", "category", "brand"]
    n = copy_into(out[cols], "fact_dispatch", cols, progress)
    return n, 0, dates, {}


def load_store_receiving(db, df, ctx, progress=None, replace=True):
    valid = _valid_stores(ctx, db)
    wh = v_str(df[_col(df, "warehouse id")], lower=True)
    dt = v_date(df[_col(df, "invoice date")])

    n_in = len(df)

    # Some exports pad the file with entirely blank rows - the 11-Aug export
    # carried 233k of them, 94% of the file. They are not lost data (no
    # status, no quantities, no timestamps) but they must be reported
    # separately from genuine out-of-network rows, or a broken export looks
    # identical to a national one.
    blank = dt.isna() & (wh == "")
    n_blank = int(blank.sum())
    df, wh, dt = df[~blank], wh[~blank], dt[~blank]

    keep = wh.isin(valid) if valid else pd.Series(True, index=df.index)
    n_foreign = int((~keep).sum())
    df, wh, dt = df[keep], wh[keep], dt[keep]
    dropped = n_blank + n_foreign

    dates = _dates_of(dt)
    if replace:
        _replace_dates(db, FactStoreReceiving, FactStoreReceiving.invoice_date, dates)

    swap_c = _col(df, "swapped")
    out = pd.DataFrame({
        "invoice_date": dt.dt.date,
        "warehouse_id": wh,
        "invoice_id": v_str(df[_col(df, "invoice id")]),
        "fsn": v_str(df[_col(df, "fsn")]),
        "description": v_str(df[_col(df, "description")]),
        "expected_qty": v_int(df[_col(df, "expected quantity")]),
        "received_qty": v_int(df[_col(df, "received quantity")]),
        "damaged_qty": v_int(df[_col(df, "damaged quantity")]),
        "scanning_issue_qty": v_int(df[_col(df, "scanning issue quantity")]),
        "excess_qty": v_int(df[_col(df, "excess quantity")]),
        "returned_qty": v_int(df[_col(df, "returned quantity")]),
        "swapped_qty": v_int(df[swap_c]) if swap_c else 0,
        "status": v_str(df[_col(df, "status")]),
        "uploaded_at": v_str(df[_col(df, "uploaded at")]),
    })
    out = _attach_cat_brand(out, ctx, db)
    cols = ["invoice_date", "warehouse_id", "invoice_id", "fsn", "description",
            "expected_qty", "received_qty", "damaged_qty", "scanning_issue_qty",
            "excess_qty", "returned_qty", "swapped_qty", "status", "uploaded_at",
            "category", "brand"]
    n = copy_into(out[cols], "fact_store_receiving", cols, progress)
    return n, dropped, dates, {"n_in": n_in, "n_blank": n_blank,
                               "n_foreign": n_foreign}


def load_wh_receiving(db, df, ctx, progress=None, replace=True):
    dt = v_date(df[_col(df, "date")])
    dates = _dates_of(dt)
    if replace:
        _replace_dates(db, FactWarehouseReceiving, FactWarehouseReceiving.date, dates)

    exp_c, cat_c, sh_c = _col(df, "expiry"), _col(df, "category"), _col(df, "short")
    fsn = v_str(df[_col(df, "fsn")])
    out = pd.DataFrame({
        "date": dt.dt.date,
        "ean": v_str(df[_col(df, "ean")], maxlen=20),
        "fsn": fsn,
        "product": v_str(df[_col(df, "product")]),
        "brand": v_str(df[_col(df, "brand")]),
        "category": df[cat_c].map(canon_category) if cat_c else v_category_from_fsn(fsn),
        "po_qty": v_int(df[_col(df, "po qty")]),
        "received_qty": v_int(df[_col(df, "received")]),
        "expiry_date": v_date(df[exp_c]).dt.date if exp_c else None,
        "short_qty": v_int(df[sh_c]) if sh_c else 0,
    })
    cols = list(out.columns)
    n = copy_into(out, "fact_wh_receiving", cols, progress)
    return n, 0, dates, {}


def load_rejects(db, df, ctx, progress=None, replace=True):
    dt = v_date(df[_col(df, "date")])
    dates = _dates_of(dt)
    if replace:
        _replace_dates(db, FactReject, FactReject.date, dates)

    fsn_c, qty_c = _col(df, "fsn"), _col(df, "qty")
    store_c, veh_c, exp_c = _col(df, "store"), _col(df, "vehicle"), _col(df, "expiry")
    cat_c = _col(df, "category")

    mask = df[fsn_c].notna()
    df, dt = df[mask], dt[mask]

    qty, corrupted = fix_excel_date_qty_vec(df[qty_c])
    fsn = v_str(df[fsn_c])
    out = pd.DataFrame({
        "date": dt.dt.date,
        "ean": v_str(df[_col(df, "ean")], maxlen=20),
        "fsn": fsn,
        "product": v_str(df[_col(df, "product")]),
        "brand": v_str(df[_col(df, "brand")]),
        "category": df[cat_c].map(canon_category) if cat_c else v_category_from_fsn(fsn),
        "qty": qty,
        "qty_was_corrupted": corrupted,
        "reason": v_str(df[_col(df, "reason")]),
        "expiry": v_date(df[exp_c]).dt.date if exp_c else None,
        "warehouse_id": v_str(df[store_c], lower=True) if store_c else None,
        "vehicle_number": v_str(df[veh_c]) if veh_c else None,
    })
    cols = list(out.columns)
    n = copy_into(out, "fact_rejects", cols, progress)

    has_store = bool(store_c) and not out["warehouse_id"].eq("").all()
    return n, 0, dates, {"n_corrupt": int(corrupted.sum()),
                         "store_attributed": 1 if has_store else 0}


def load_route(db, df, ctx, progress=None, replace=True):
    valid = _valid_stores(ctx, db)
    dt = v_date(df[_col(df, "date")])
    dates = _dates_of(dt)
    if replace:
        _replace_dates(db, FactRoute, FactRoute.date, dates)

    store_c = _col(df, "store no") or _col(df, "store")
    raw_store = v_str(df[store_c], lower=True)
    co_c, ci_c = _col(df, "crate out"), _col(df, "crate in")
    st_c, en_c = _col(df, "start"), _col(df, "end")
    sno_c, drv_c, veh_c, rm_c = (_col(df, "sno"), _col(df, "driver"),
                                 _col(df, "vehicle"), _col(df, "remark"))

    out = pd.DataFrame({
        "date": dt.dt.date,
        "stop_seq": v_int(df[sno_c]) if sno_c else 0,
        "warehouse_id": raw_store.where(raw_store.isin(valid)),
        "store_name": v_str(df[store_c]),
        "driver": v_str(df[drv_c]) if drv_c else None,
        "vehicle_no": v_str(df[veh_c]) if veh_c else None,
        "out_time": v_str(df[st_c]) if st_c else None,
        "in_time": v_str(df[en_c]) if en_c else None,
        "crate_out": v_int_null(df[co_c]) if co_c else None,
        "crate_in": v_int_null(df[ci_c]) if ci_c else None,
        "remark": v_str(df[rm_c]) if rm_c else None,
    })
    cols = list(out.columns)
    n = copy_into(out, "fact_route", cols, progress)
    return n, 0, dates, {}


def load_indent(db, df, ctx, progress=None, replace=True):
    df = df.loc[:, ~df.columns.astype(str).str.startswith("Unnamed")]
    dt = v_date(df[_col(df, "indent_date") or _col(df, "indent")])
    dates = _dates_of(dt)
    if replace:
        _replace_dates(db, FactIndent, FactIndent.indent_date, dates)

    frq, pod_c = _col(df, "final received"), _col(df, "po date")
    out = pd.DataFrame({
        "indent_date": dt.dt.date,
        "po_date": v_date(df[pod_c]).dt.date if pod_c else None,
        "brand": v_str(df[_col(df, "brand")]),
        "fsn": v_str(df[_col(df, "fsn")]),
        "po_qty": v_int(df[_col(df, "po qty")]),
        "vertical": v_str(df[_col(df, "vertical")]),
        "title": v_str(df[_col(df, "title")]),
        "final_received_qty": v_int_null(df[frq]) if frq else None,
    })
    cols = list(out.columns)
    n = copy_into(out, "fact_indent", cols, progress)

    # Counted as rows-with-data so a partially-closed loop across chunks
    # doesn't read as fully open.
    closed = int(df[frq].notna().sum()) if frq else 0
    return n, 0, dates, {"indent_rows": len(df), "indent_closed": closed}


LOADERS = {
    "store_master": load_store_master,
    "product_master": load_product_master,
    "batching": load_batching,
    "store_receiving": load_store_receiving,
    "wh_receiving": load_wh_receiving,
    "rejects": load_rejects,
    "route": load_route,
    "indent": load_indent,
}

# Which column carries the date, and which table/column to clear for it.
# Drives the cheap up-front date scan in chunked mode.
DATE_SPEC = {
    "batching": (("cutoff",), FactDispatch, FactDispatch.dispatch_date),
    "store_receiving": (("invoice date",), FactStoreReceiving,
                        FactStoreReceiving.invoice_date),
    "wh_receiving": (("date",), FactWarehouseReceiving,
                     FactWarehouseReceiving.date),
    "rejects": (("date",), FactReject, FactReject.date),
    "route": (("date",), FactRoute, FactRoute.date),
    "indent": (("indent",), FactIndent, FactIndent.indent_date),
}


# ---------------- notes ----------------
def summarise(ftype, stats, dates, loaded):
    """Turn accumulated counters into operator-facing notes, once."""
    notes = []
    if ftype == "store_master":
        notes.append("Store master replaced in full.")
    elif ftype == "product_master":
        notes.append(f"Category casing normalised; {stats.get('dupes', 0)} "
                     f"duplicate FSN rows collapsed.")
    elif ftype == "batching":
        notes.append(f"Dispatch loaded for {len(dates)} date(s).")
    elif ftype == "store_receiving":
        notes.append(f"Loaded {len(dates)} date(s).")
        n_in, n_blank = stats.get("n_in", 0), stats.get("n_blank", 0)
        n_foreign = stats.get("n_foreign", 0)
        if n_blank and n_in:
            notes.append(
                f"{n_blank:,} of {n_in:,} rows ({n_blank/n_in:.0%}) were completely "
                f"blank padding from the export - no store, date, status or "
                f"quantities. Ignored. If you expected more days of data, the export "
                f"likely didn't run over the full date range.")
        if n_foreign:
            notes.append(
                f"{n_foreign:,} rows belonged to darkstores outside your 48-store "
                f"network and were excluded - Flipkart's export is national.")
    elif ftype == "wh_receiving":
        notes.append("Category casing normalised.")
    elif ftype == "rejects":
        n_corrupt = stats.get("n_corrupt", 0)
        if n_corrupt:
            notes.append(
                f"{n_corrupt:,} of {loaded:,} QTY values had been converted to "
                f"dates by Excel. Original integers recovered - set that column "
                f"to Number at source to stop this recurring.")
        if not stats.get("store_attributed"):
            notes.append("No store attribution in this file - rejects can't be "
                         "traced to a specific darkstore.")
    elif ftype == "route":
        notes.append("Crate counts are per-vehicle, per-trip - they narrow a "
                     "swap to a route, not to a specific stop.")
    elif ftype == "indent":
        if stats.get("indent_rows") and not stats.get("indent_closed"):
            notes.append("'Final Received Qty' is empty for every row - the indent "
                         "loop isn't being closed operationally.")
    return notes


def scan_dates(path, encoding, date_needles, sample_cols):
    """
    Read ONLY the date column to learn which dates the file covers, so the
    idempotent delete can happen once before any chunk is inserted. Reading
    one column of a large file costs a fraction of reading all of it.
    """
    probe = pd.DataFrame(columns=sample_cols)
    col = _col(probe, *date_needles)
    if col is None:
        return []
    dates = set()
    try:
        for chunk in iter_chunks(path, encoding, usecols=[col]):
            dates.update(_dates_of(v_date(chunk[col])))
    except (ValueError, KeyError):
        return []
    return sorted(dates)


# --------------------------------------------------------------------------
# Driver
# --------------------------------------------------------------------------
def ingest_path(filename, path, user_email, forced_type=None, progress=None):
    """
    Import a file already spooled to disk. Memory use is bounded by
    CHUNK_ROWS regardless of how large the file is.
    """
    db = SessionLocal()
    tmpdir = tempfile.mkdtemp(prefix="ingest_")
    converted = None
    ctx = {}
    try:
        csv_path, was_converted = to_csv_path(filename, path, tmpdir)
        converted = csv_path if was_converted else None
        encoding = detect_encoding(csv_path)

        sample = read_sample(csv_path, encoding)
        ftype = forced_type or detect_type(sample)[0]
        if not ftype:
            raise ValueError("Could not recognise this file's columns. "
                             "Pick the type manually.")

        loader = LOADERS[ftype]
        total_loaded = total_dropped = rows_in = 0
        all_dates, stats = set(), {}

        if ftype in MASTER_TYPES:
            # Small, full-replace, and needs cross-file de-duplication.
            df = read_whole(csv_path, encoding)
            rows_in = len(df)
            total_loaded, total_dropped, dates, stats = loader(
                db, df, ctx, progress, replace=True)
            all_dates.update(dates)
            del df
        else:
            # Clear the affected dates ONCE, then stream the file in.
            needles, model, date_col = DATE_SPEC[ftype]
            file_dates = scan_dates(csv_path, encoding, needles,
                                    list(sample.columns))
            _replace_dates(db, model, date_col, file_dates)

            for chunk in iter_chunks(csv_path, encoding):
                rows_in += len(chunk)
                n, dropped, dates, st = loader(db, chunk, ctx, None,
                                               replace=False)
                total_loaded += n
                total_dropped += dropped
                all_dates.update(dates)
                for k, v in st.items():
                    stats[k] = stats.get(k, 0) + v
                if progress:
                    progress(rows_in, None)
                del chunk

        dates_sorted = sorted(all_dates)
        notes = summarise(ftype, stats, dates_sorted, total_loaded)

        db.add(UploadLog(
            uploaded_by=user_email, filename=filename, file_type=ftype,
            dates_covered=", ".join(str(d) for d in dates_sorted) if dates_sorted else "-",
            rows_in_source=rows_in, rows_loaded=total_loaded,
            rows_dropped=total_dropped, notes=" ".join(notes), status="ok"))
        db.commit()
        return {"ok": True, "type": ftype, "rows_in": rows_in,
                "rows_loaded": total_loaded, "rows_dropped": total_dropped,
                "dates": [str(d) for d in dates_sorted], "notes": notes,
                "filename": filename}
    except Exception as e:
        db.rollback()
        try:
            db.add(UploadLog(uploaded_by=user_email, filename=filename,
                             file_type=forced_type or "unknown", rows_in_source=0,
                             rows_loaded=0, rows_dropped=0,
                             notes=str(e)[:4000], status="error"))
            db.commit()
        except Exception:
            db.rollback()
        return {"ok": False, "error": str(e), "filename": filename,
                "type": forced_type or "unknown", "rows_in": 0,
                "rows_loaded": 0, "rows_dropped": 0, "dates": [], "notes": []}
    finally:
        db.close()
        if converted and os.path.exists(converted):
            try:
                os.unlink(converted)
            except OSError:
                pass
        try:
            os.rmdir(tmpdir)
        except OSError:
            pass


def ingest_file(filename, content, user_email, forced_type=None, progress=None):
    """
    Backwards-compatible wrapper for callers that still hold bytes
    (seed_test.py, build_preview.py). Spools to disk, then uses the
    streaming path. New code should call ingest_path directly.
    """
    fd, tmp = tempfile.mkstemp(suffix=Path(filename).suffix or ".csv")
    try:
        with os.fdopen(fd, "wb") as fh:
            fh.write(content)
        return ingest_path(filename, tmp, user_email, forced_type, progress)
    finally:
        if os.path.exists(tmp):
            try:
                os.unlink(tmp)
            except OSError:
                pass
